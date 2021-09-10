# Time will start for the calculation of execution time
import time
start_time = time.time()

###############################################################################################################

# Check all modules are present
import subprocess
import sys

modules = [
    "requests",
    "xlsxwriter",
    "datetime",
    "openpyxl"
]
for check_module in modules:
    subprocess.check_call([sys.executable, "-m", "pip", "install", check_module])
print("----Good To Go----")

###############################################################################################################

# Beginning of the program
import os
import requests
import json
import xlsxwriter
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint

###############################################################################################################

req = requests.get('https://www.mohfw.gov.in/data/datanew.json')  # Get data from the server
req.raise_for_status()  # Check if there are any errors
data = json.loads(req.text)  # Load json

# Insert data accordingly into the variable table
table = []
for i in range(37):
    rule = {}
    if i == 18:
        continue
    # Add all info into rule
    if data[i]['sno'] == '11111':
        rule["S.No."] = ""
    else:
        rule["S.No."] = i + 1
    if data[i]['state_name'] == '':
        rule["Name of State / UT"] = 'Total'
    else:
        rule["Name of State / UT"] = data[i]['state_name']
    rule["Active Cases"] = data[i]['new_active']
    rule["New Active"] = int(data[i]['new_positive']) - int(data[i]['positive'])
    rule["Cured/Discharged/Migrated"] = data[i]['new_cured']
    rule["New Cured"] = int(data[i]['new_cured']) - int(data[i]['cured'])
    rule["Differences"] = rule["New Cured"] - rule["New Active"]
    try:
        rule[" % Differences"] = round(rule["Differences"] / rule["New Cured"] * 100, 1)
    except:
        rule[" % Differences"] =  0
    rule["Deaths"] = data[i]['new_death']
    rule["New Deaths"] = int(data[i]['new_death']) - int(data[i]['death'])
    rule["Total Confirmed cases"] = data[i]['new_positive']
    # Recovery rate calculation
    if rule["Total Confirmed cases"] == '0':
        rule["Recovery Rate ( % )"] = 0
    else:
        recovery = int(rule["Cured/Discharged/Migrated"]) / int(rule["Total Confirmed cases"])
        rule["Recovery Rate ( % )"] = round(recovery * 100, 1)
    table.append(rule)

###############################################################################################################

# Open and Create an Excel workbook
workbook = xlsxwriter.Workbook('Status_Rate_4.0.xlsx', {'default_date_format': 'dd/mm/yy'})
# By default worksheet names in the spreadsheet will be Sheet1, Sheet2 etc., but we can also specify a name:
# worksheet1 = workbook.add_worksheet()        # Defaults to Sheet1.
# worksheet2 = workbook.add_worksheet('Data')  # Data.
worksheet = workbook.add_worksheet()

# Set up a format
headers_font = workbook.add_format({'bold': True, 'font_size': 18})
name_font = workbook.add_format({'font_size': 14, 'font_color': 'gray', 'align': 'left'})
value_font = workbook.add_format({'font_size': 14, 'font_color': 'gray', 'align': 'right'})
active_death_font = workbook.add_format({'bold': True, 'font_size': 12, 'font_color': 'red', 'align': 'right'})
cured_font = workbook.add_format({'bold': True, 'font_size': 12, 'font_color': 'green', 'align': 'right'})
# Font of the date with alignment
date_font = workbook.add_format({'bold': True, 'font_size': 12, 'align': 'center'})
# Font of the date and time
time_font = workbook.add_format(
    {'font_size': 12, 'font_color': 'gray', 'align': 'center', 'num_format': 'dd/mm/yyyy hh:mm AM/PM'})
# Add a format. Light red fill with dark red text.
red_format = workbook.add_format({'font_size': 14, 'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
# Add a format. Green fill with dark green text.
green_format = workbook.add_format({'font_size': 14, 'bg_color': '#C6EFCE', 'font_color': '#006100'})
# Add a format. Green fill with high value dark green text.
high_green_format = workbook.add_format({'bg_color': '#A4DE02', 'font_color': '#006100'})

# Headings of all columns
headers = [
    "S.No.",
    "Name of State / UT",
    "Active Cases",
    "New Active",
    "Cured/Discharged/Migrated",
    "New Cured",
    "Differences",
    " % Differences",
    "Deaths",
    "New Deaths",
    "Total Confirmed cases",
    "Recovery Rate ( % )"
]

# We use the worksheet object to write headers via the write() method:
# worksheet.write(row, col, some_data)
for col_num, header in enumerate(headers):
    worksheet.write(0, col_num, header, headers_font)

# Write a date.
date_time = datetime.datetime.now()
worksheet.write(0, 14, 'Date:', date_font)
worksheet.set_column('O:Q', 22)
worksheet.write_datetime('P1', date_time, time_font)  # 07/09/20 08:35:34 AM (dd/mm/yyyy hh:mm AM/PM)

# Reduced value from total and Font format of reduction/addition total
reduced_font = workbook.add_format({'bold': True, 'font_size': 14, 'font_color': 'green', 'align': 'center'})
add_font = workbook.add_format({'bold': True, 'font_size': 14, 'font_color': 'red', 'align': 'center'})
worksheet.write('O2', 'Total Difference:', date_font)
dif = table[35]["New Cured"] + table[35]["New Deaths"] - table[35]["New Active"]
if dif > 0:
    worksheet.write('P2', str(dif) + ' Cases Cured', reduced_font)
else:
    worksheet.write('P2', str(abs(dif)) + ' Cases Added', add_font)

# Save the data into the excel sheet
for i in range(36):
    for col_num, cell_data in enumerate(table[i].values()):
        if col_num == 1:
            worksheet.write(i + 1, col_num, cell_data, name_font)
        elif col_num == 3:
            worksheet.write(i + 1, col_num, cell_data, active_death_font)
        elif col_num == 5:
            worksheet.write(i + 1, col_num, cell_data, cured_font)
        elif col_num == 6:
            if cell_data > 0:
                worksheet.write(i + 1, col_num, cell_data, green_format)
            else:
                worksheet.write(i + 1, col_num, cell_data, red_format)
        elif col_num == 7:
            if cell_data > 0:
                worksheet.write(i + 1, col_num, cell_data, green_format)
            else:
                worksheet.write(i + 1, col_num, cell_data, red_format)
        elif col_num == 9:
            worksheet.write(i + 1, col_num, cell_data, active_death_font)
        elif col_num == 11:
            worksheet.write(i + 1, col_num, cell_data, value_font)
        else:
            worksheet.write(i + 1, col_num, cell_data, value_font)

# Write conditional formats over the same range for the color of cells
worksheet.conditional_format('L2:L37', {'type': 'cell',
                                        'criteria': 'between',
                                        'minimum': 70,
                                        'maximum': 85,
                                        'format': green_format})

worksheet.conditional_format('L2:L37', {'type': 'cell',
                                        'criteria': '>=',
                                        'value': 85,
                                        'format': high_green_format})

worksheet.conditional_format('L2:L37', {'type': 'cell',
                                        'criteria': '<',
                                        'value': 70,
                                        'format': red_format})
# Close the workbook
workbook.close()

###############################################################################################################

# Read the workbook
wb = load_workbook('Status_Rate_4.0.xlsx')
ws = wb['Sheet1']


# Create a Class for Multiple charts
class charting:

    def __init__(self, ch):  # Create a new chart object according to the ch condition .
        if ch == 'col':
            self.chart = BarChart()
        else:
            self.chart = LineChart()

    def add_series(self, cond=None):  # Add a series to the chart.
        if cond == 1:
            # Add a series to the chart 1.
            self.data = Reference(ws, min_col=12, min_row=1, max_row=37, max_col=12)
        else:
            # Add a series to the chart 2
            self.data = Reference(ws, min_col=8, min_row=1, max_row=37, max_col=8)
        self.cats = Reference(ws, min_col=2, min_row=2, max_row=37)
        self.chart.add_data(self.data, titles_from_data=True)
        self.chart.set_categories(self.cats)
        self.chart.shape = 4

    def chart_axis(self, tname=None, yname=None, xname=None, ty=None, led=None):  # Customize the chart axis
        self.chart.style = 10
        self.chart.type = ty  # "col"
        self.chart.legend = led  # None
        self.chart.title = tname  # 'Chart of Recovery Rate'
        self.chart.y_axis.title = yname  # 'Recovery Rate ( % )'
        self.chart.x_axis.title = xname  # 'Name of State / UT'
        self.chart.x_axis.tickLblSkip = 1  # list all name in x-axis

    def add_color(self, cond=None):  # Add colors into all bars of chart
        s = self.chart.series[0]
        if cond == 1:  # Add color in chart 1 for the recovery rate
            for num in range(36):
                if table[num]["Name of State / UT"] == "Total":
                    pt = DataPoint(idx=num)
                    pt.graphicalProperties.solidFill = "008000"  # Very Dark Color of Green
                    s.dPt.append(pt)
                    continue
                if table[num]["Recovery Rate ( % )"] >= 85:
                    pt = DataPoint(idx=num)
                    pt.graphicalProperties.solidFill = "00C000"  # Dark Green Color
                    s.dPt.append(pt)
                elif 85 > table[num]["Recovery Rate ( % )"] >= 70:
                    pt = DataPoint(idx=num)
                    pt.graphicalProperties.solidFill = "C6EFCE"  # Green Color
                    s.dPt.append(pt)
                else:
                    pt = DataPoint(idx=num)
                    pt.graphicalProperties.solidFill = "FF0000"  # Red Color
                    s.dPt.append(pt)
        elif cond == 2:
            for nm in range(36):  # Add color in chart 2 for percentage difference
                if table[nm][" % Differences"] >= 50:
                    pt = DataPoint(idx=nm)
                    pt.graphicalProperties.solidFill = "00C000"  # Dark Green Color
                    s.dPt.append(pt)
                elif 50 > table[nm][" % Differences"] >= 0:
                    pt = DataPoint(idx=nm)
                    pt.graphicalProperties.solidFill = "A4DE02"  # Green Color
                    s.dPt.append(pt)
                else:
                    pt = DataPoint(idx=nm)
                    pt.graphicalProperties.solidFill = "FF0000"  # Red Color
                    s.dPt.append(pt)
        else:
            s.graphicalProperties.line.solidFill = "00AAAA"
            s.graphicalProperties.line.width = 20000

    def location_insert(self, lct):  # Insert the chart into the worksheet
        ws.add_chart(self.chart, lct)


###############################################################################################################

# Chart 1 is the Bar Graph of the Recovery Rate
chart1 = charting('col')
chart1.add_series(1)
chart1.chart_axis('Chart of Recovery Rate', 'Recovery Rate ( % )', 'Name of State / UT', 'col')
chart1.add_color(1)
chart1.location_insert('M3')

# Chart 2 is the Bar Graph of the percentage of daily difference between cured and active
chart2 = charting('col')
chart2.add_series(2)
chart2.chart_axis('Chart of Cured/Active Cases Difference', 'Differences (%)', 'Name of State / UT', 'col')
chart2.add_color(2)
# Added Line Chart with the Bar Chart 2
chart3 = charting('line')
chart3.add_series(2)
chart3.add_color()
chart2.chart += chart3.chart
chart2.location_insert('M14')

# Print the program execution time in excel file
ws["Q1"] = "Execution Time:"
ws["Q1"].font = Font(bold=True, size="12")
ws["Q2"] = " " + str(round(time.time() - start_time, 4)) + " seconds"
ws["Q2"].font = Font(size="12")

# Save the Graph into the Status_Rate_4.0.xlsx file
wb.save('Status_Rate_4.0.xlsx')
# Open file automatically
os.startfile('Status_Rate_4.0.xlsx')